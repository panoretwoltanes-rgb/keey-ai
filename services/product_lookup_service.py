"""产品精确查找服务 - 唯一数据库记录读取"""
import openpyxl

DB_PATH = r"D:\0000AAA报价助手\1.2026企一产品报价表(最新）.xlsx"


def exact_db_lookup(model_keyword: str, qty: int = 1, force_wb: bool = False) -> dict:
    """在产品库中按型号精确查找，返回唯一数据库记录。
    
    查找优先级：
      1. D 列（型号）精确匹配
      2. D 列前缀匹配（如 "QY-TH06055S" 匹配 "QY-TH06055S 白+白"）
      3. C 列（货号）精确匹配
      4. C 列前缀匹配
    
    返回完整产品数据：货号/型号/参数/开孔/单位/单价/db_sheet/db_row。
    """
    if not model_keyword:
        return {}
    lookup_key = model_keyword.strip()
    # 预埋订单强制使用 WB 系列（不允许返回 TH 数据）
    if force_wb and lookup_key.startswith("QY-TH"):
        lookup_key = lookup_key.replace("QY-TH", "QY-WB", 1)
        # 颜色格式映射：白+白→白 / 黑+黑→黑 / 镍+镍→镍
        parts = lookup_key.split(maxsplit=1)
        if len(parts) > 1 and parts[1] in {"白+白", "黑+黑", "镍+镍"}:
            lookup_key = f"{parts[0]} {parts[1].split('+')[0]}"
        print(f"[DB LOOKUP] 预埋强制 WB: {model_keyword} -> {lookup_key}")
    kw = lookup_key.lower()
    wb = openpyxl.load_workbook(DB_PATH, data_only=True)
    best = None
    try:
        for sname in wb.sheetnames:
            if "恒流" in sname:
                continue
            ws = wb[sname]
            for row_idx in range(3, (ws.max_row or 3) + 1):
                c3 = str(ws.cell(row=row_idx, column=3).value or "").strip()
                c4 = str(ws.cell(row=row_idx, column=4).value or "").strip()
                c5 = str(ws.cell(row=row_idx, column=5).value or "").strip()
                c6 = str(ws.cell(row=row_idx, column=6).value or "").strip()
                c8 = str(ws.cell(row=row_idx, column=8).value or "").strip()
                c9 = ws.cell(row=row_idx, column=9).value or 0
                c4l = c4.lower()
                c3l = c3.lower()
                matched = False
                if c4l == kw:
                    matched = True
                elif c4l.startswith(kw):
                    matched = True
                elif c3l == kw:
                    matched = True
                elif c3l.startswith(kw):
                    matched = True
                if matched:
                    best = {
                        "货号": c3, "型号": c4, "参数": c5,
                        "开孔": c6, "单位": c8, "单价": c9,
                        "数量": qty,
                        "db_sheet": sname, "db_row": row_idx,
                    }
                    break
            if best:
                break
        # 尝试 QY- 前缀补全（如 ZH031MF-1 → QY-ZH031MF-1）
        if not best and not kw.startswith("qy-"):
            kw2 = ("QY-" + lookup_key).lower()
            for sname in wb.sheetnames:
                if "恒流" in sname:
                    continue
                ws = wb[sname]
                for row_idx in range(3, (ws.max_row or 3) + 1):
                    c3 = str(ws.cell(row=row_idx, column=3).value or "").strip()
                    c4 = str(ws.cell(row=row_idx, column=4).value or "").strip()
                    c5 = str(ws.cell(row=row_idx, column=5).value or "").strip()
                    c6 = str(ws.cell(row=row_idx, column=6).value or "").strip()
                    c8 = str(ws.cell(row=row_idx, column=8).value or "").strip()
                    c9 = ws.cell(row=row_idx, column=9).value or 0
                    c4l = c4.lower()
                    c3l = c3.lower()
                    if c4l == kw2 or c4l.startswith(kw2) or c3l == kw2 or c3l.startswith(kw2):
                        best = {
                            "货号": c3, "型号": c4, "参数": c5,
                            "开孔": c6, "单位": c8, "单价": c9,
                            "数量": qty,
                            "db_sheet": sname, "db_row": row_idx,
                        }
                        print(f"[精确查库] QY-补全: {lookup_key} -> QY-{lookup_key}")
                        break
                if best:
                    break
        if not best:
            print(f"[精确查库] 未找到: {model_keyword}")
    finally:
        wb.close()
    return best or {}
