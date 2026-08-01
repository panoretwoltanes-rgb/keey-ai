"""Excel 生成服务 - 调用 V1 正式报价引擎 + 产品匹配验证"""
import sys, os, shutil, importlib.util, time, re, json
import openpyxl
from datetime import date
from services.product_lookup_service import exact_db_lookup

_V1_SCRIPTS = r"E:\新建文件夹 (2)\灯具报价\KEEY报价助手_V1.0_Stable_RESTORE\scripts"
_V2_OUT = r"E:\新建文件夹 (2)\灯具报价\KEEY_AI_Quote\output"

# 加载 V1 引擎
def _load_v1_engine():
    _old_path = sys.path.copy()
    _old_mods = {}
    for n in list(sys.modules.keys()):
        if n in ("config", "product_resolver", "template_handler",
                 "image_handler", "rich_text_handler", "quote_generator"):
            _old_mods[n] = sys.modules[n]
    _stdlib = [p for p in sys.path if p.lower().startswith(os.path.dirname(os.__file__).lower())]
    sys.path = _stdlib + [_V1_SCRIPTS]
    for n in list(sys.modules.keys()):
        if n in ("config", "product_resolver", "template_handler",
                 "image_handler", "rich_text_handler", "quote_generator"):
            del sys.modules[n]
    spec = importlib.util.spec_from_file_location("v1_quote_gen",
        os.path.join(_V1_SCRIPTS, "quote_generator.py"))
    mod = importlib.util.module_from_spec(spec)
    old_cwd = os.getcwd()
    os.chdir(_V1_SCRIPTS)
    try:
        spec.loader.exec_module(mod)
    finally:
        os.chdir(old_cwd)
    # 加载 V1 resolver
    rspec = importlib.util.spec_from_file_location("v1_resolver",
        os.path.join(_V1_SCRIPTS, "product_resolver.py"))
    rmod = importlib.util.module_from_spec(rspec)
    os.chdir(_V1_SCRIPTS)
    try:
        rspec.loader.exec_module(rmod)
    finally:
        os.chdir(old_cwd)
    sys.path = _old_path
    for n, m in _old_mods.items():
        if n not in sys.modules:
            sys.modules[n] = m
    return mod.generate_quote, rmod.resolve_products

_generate_quote, _resolve_products = _load_v1_engine()


def merge_product_params(db_params: str, order_item: dict) -> str:
    """统一产品参数合并入口（全流程唯一参数修改函数）。
    
    负责：
      ① 色温替换（如 "3000K/4000K" → "4000K"）
      ② 光束角替换（如 "24°（15/36/50°）" → "24°"）
      ③ 用户自定义参数
    
    安装方式不再在此修改：预埋产品在匹配阶段直接使用 WB 系列数据，
    数据库参数天然正确。其它参数行原样保留。
    """
    if not db_params:
        return db_params
    cct = order_item.get("cct", "")
    beam = order_item.get("beam", "")

    lines = db_params.split("\n")
    new_lines = []
    cct_done = False
    beam_done = False

    for line in lines:
        # ① 色温替换
        if cct and "色温" in line and not cct_done:
            m = re.match(r"^(\d+\.\s*)", line)
            prefix = m.group(1) if m else ""
            new_lines.append(f"{prefix}色温：{cct}")
            cct_done = True
            continue
        # ② 光束角替换
        if beam and "光束角" in line and not beam_done:
            m = re.match(r"^(\d+\.\s*)", line)
            prefix = m.group(1) if m else ""
            new_lines.append(f"{prefix}光束角：{beam.replace('度', '°')}")
            beam_done = True
            continue
        new_lines.append(line)

    # 数据库缺失对应行时补充
    if cct and not cct_done:
        new_lines.append(f"色温：{cct}")
    if beam and not beam_done:
        new_lines.append(f"光束角：{beam.replace('度', '°')}")

    return "\n".join(new_lines)

def _fill_header_info(ws, customer: str, project: str):
    """填充报价单头部的客户/项目信息。"""
    filled_customer = False
    filled_project = False
    for row in range(1, 7):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            val = str(cell.value or "").strip()
            if not val:
                continue
            if not filled_customer and "客户" in val and customer:
                clean = val.rstrip().rstrip("：:/ ")
                cell.value = f"{clean}：{customer}"
                filled_customer = True
                print(f"[头部] 客户写入 {cell.coordinate}: {cell.value}")
                continue
            if not filled_project and "项目" in val and project:
                clean = val.rstrip().rstrip("：:/ ")
                cell.value = f"{clean}：{project}"
                filled_project = True
                print(f"[头部] 项目写入 {cell.coordinate}: {cell.value}")
                continue
            if val == "/" and not filled_customer and customer:
                for pcol in range(col - 1, 0, -1):
                    pval = str(ws.cell(row=row, column=pcol).value or "")
                    if "客户" in pval:
                        cell.value = customer
                        filled_customer = True
                        print(f"[头部] 客户写入 {cell.coordinate}: {customer}")
                        break
                continue
            if val == "/" and not filled_project and project:
                for pcol in range(col - 1, 0, -1):
                    pval = str(ws.cell(row=row, column=pcol).value or "")
                    if "项目" in pval:
                        cell.value = project
                        filled_project = True
                        print(f"[头部] 项目写入 {cell.coordinate}: {project}")
                        break
                continue
    return ws


def generate_excel(order_data: dict) -> str:
    customer = order_data.get("customer", "客户")
    date_str = date.today().strftime("%Y-%m-%d")
    products = order_data.get("products", [])
    if not products:
        raise ValueError("订单无产品")
    _force_wb = "预埋" in order_data.get("_accessory_directives", [])

    # 第一步：构建 v1_products（优先使用订单已携带的 _db_record，不再重复查库）
    v1_products = []
    for p in products:
        # 订单已携带数据库记录：直接使用，不再查库
        db_prod = p.get("_db_record") or {}
        if not db_prod:
            # 无 _db_record（如配件）：按型号查库
            model = p.get("model", "")
            color = p.get("color", "")
            qty = p.get("quantity", 0)
            search_keyword = f"{model} {color}" if (color and color not in model) else model
            db_prod = exact_db_lookup(search_keyword, qty, force_wb=_force_wb)
        if db_prod:
            print("======== DB LOOKUP ========")
            print(f"Model: {db_prod.get('型号', p.get('model',''))}")
            print(f"Found: {db_prod.get('型号','')}")
            print(f"Sheet: {db_prod.get('db_sheet','')}")
            print(f"Row: {db_prod.get('db_row','')}")
            print("===========================")
            # 统一参数合并：唯一参数处理入口
            order_item = dict(p)
            db_params = db_prod.get("参数", "") or ""
            db_prod["参数"] = merge_product_params(db_params, order_item)
            print(f"[参数合并] 最终参数:\n{db_prod['参数']}")
            v1_products.append(db_prod)
            continue

        # 查库失败：直接报错
        print(f"[LOOKUP ERROR] Model: {p.get('model','')}")
        raise ValueError(f"产品未找到: {p.get('model','')}")

    # 第二步：直接调用 V1 模板函数生成 Excel（不经过 generate_quote 避免二次覆盖）
    import config as v1_cfg
    import template_handler as tpl
    import image_handler as img_h
    import rich_text_handler as rt

    print(f"\n======== WRITE EXCEL ========")
    for vp in v1_products:
        print(json.dumps(vp, ensure_ascii=False, indent=2))
    print("============================")

    wb_db = openpyxl.load_workbook(v1_cfg.PRODUCT_DB_PATH)
    product_count = len(v1_products)
    v1_products = img_h.copy_multiple_images(wb_db, v1_products)

    wb_tpl = openpyxl.load_workbook(v1_cfg.TEMPLATE_PATH)
    ws_tpl = wb_tpl.active
    logo = tpl.preserve_logo(ws_tpl)
    _fill_header_info(ws_tpl, customer, order_data.get("project", ""))
    tpl.rebuild_data_area(ws_tpl, product_count)

    footer_start = 7 + product_count
    tpl.build_footer(ws_tpl, footer_start, product_count)
    tpl.update_date(ws_tpl, date_str)

    print("======== BEFORE FILL ========")
    for _vp in v1_products:
        _debug = {k: v for k, v in _vp.items() if k != "img"}
        print(json.dumps(_debug, ensure_ascii=False, indent=2))
    print("=============================")
    tpl.fill_product_data(ws_tpl, v1_products)
    print("======== BEFORE IMAGE ========")
    for _vp in v1_products:
        print(json.dumps({k:v for k,v in _vp.items() if k != 'img'}, ensure_ascii=False, indent=2))
    print("==============================")
    tpl.add_images(ws_tpl, logo, v1_products)
    print(f"  图片共计: {len(ws_tpl._images)}张")
    tpl.restore_page_setup(ws_tpl)

    output_name = f"{date_str}-{customer}灯具报价表.xlsx"
    tmpfile = os.path.join(v1_cfg.TEMP_DIR, f"tmp_{customer}_{int(time.time())}.xlsx")
    if os.path.exists(tmpfile):
        os.remove(tmpfile)
    wb_tpl.save(tmpfile)
    rt.apply_red_text(tmpfile)

    v1_output = os.path.join(v1_cfg.OUTPUT_DIR, output_name)
    try:
        if os.path.exists(v1_output):
            os.remove(v1_output)
        shutil.copy2(tmpfile, v1_output)
        print(f"\n[OK] {v1_output}")
    except PermissionError:
        print(f"\n[!!] 文件被锁定，使用临时路径")
        v1_output = tmpfile

    print(f"合计: {sum(p['单价'] * p['数量'] for p in v1_products):.2f}")
    for p in v1_products:
        amt = p["单价"] * p["数量"]
        print(f"  {p['型号'][:25]} x {p['数量']}{p['单位']} = {amt:>8.2f}")

    # V2 输出目录（硬编码绝对路径，不与 V1 config 冲突）
    os.makedirs(_V2_OUT, exist_ok=True)
    v2_path = os.path.join(_V2_OUT, output_name)

    print(f"[源文件] {v1_output}")
    print(f"[源文件存在] {os.path.exists(v1_output)}")
    print(f"[目标目录] {_V2_OUT}")
    print(f"[目标文件] {v2_path}")
    print(f"[源文件与目标文件是否相同] {os.path.abspath(v1_output) == os.path.abspath(v2_path)}")

    if os.path.abspath(v1_output) == os.path.abspath(v2_path):
        raise RuntimeError("源文件与目标文件相同")

    if os.path.exists(v2_path):
        try:
            os.remove(v2_path)
        except PermissionError:
            name_only, ext = os.path.splitext(output_name)
            v2_path = os.path.join(_V2_OUT, f"{name_only}_{int(time.time())}{ext}")

    shutil.copy2(v1_output, v2_path)
    print(f"[复制成功] {v2_path}")
    print(f"[目标文件存在] {os.path.exists(v2_path)}")
    return os.path.basename(v2_path)
