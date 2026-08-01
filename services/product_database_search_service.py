"""产品数据库搜索服务 - 评分排序 + 强制过滤"""
import os

DB_PATH = r"D:\0000AAA报价助手\1.2026企一产品报价表(最新）.xlsx"

EXCLUDE_WORDS = [
    "工具", "配件", "堵头", "镜片", "预埋盒",
    "驱动", "电源", "卡扣", "支架", "接头",
    "螺丝", "弹簧", "裸板", "PCB", "线路板",
    "贴合工具", "安装工具", "套管", "连接件", "型材",
]


PRODUCT_TYPES = ["筒灯", "射灯", "磁吸灯", "灯带", "洗墙灯", "轨道灯", "天花灯", "格栅灯"]


def _is_product_type(kw: str) -> bool:
    for t in PRODUCT_TYPES:
        if t in kw or kw in t:
            return True
    return False


def _load_keyword_rules() -> dict:
    import json
    rules_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "product_keyword_rules.json")
    try:
        with open(rules_path, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
            return data
    except:
        return {}


def _get_color_alias_map() -> dict:
    """从 keyword_rules JSON 加载颜色别名映射。"""
    rules = _load_keyword_rules()
    if not isinstance(rules, dict):
        return {}
    return rules.get("color_alias", {})


def _resolve_color_keyword(input_color: str) -> str:
    """将客户颜色叫法转换为产品颜色关键词。"""
    if not input_color:
        return ""
    alias_map = _get_color_alias_map()
    if input_color in alias_map:
        variants = alias_map[input_color]
        return variants[0] if variants else input_color
    return input_color


_rules_raw = _load_keyword_rules()
if isinstance(_rules_raw, dict):
    _keyword_rules = _rules_raw.get("keywords", {})
    _color_aliases = _rules_raw.get("color_alias", {})
else:
    _keyword_rules = {}
    _color_aliases = {}


def _check_keyword_rules(model: str, keyword: str, rules: dict) -> tuple:
    """检查型号是否匹配关键词规则。返回 (score, field) 或 (0, "")"""
    kw = keyword.lower().strip()
    mdl = model.lower()
    for rule_kw, rule in rules.items():
        if kw == rule_kw or (len(rule_kw) > 1 and rule_kw in kw) or (len(kw) > 1 and kw in rule_kw):
            match_any = rule.get("match_any", [])
            if any(p in mdl for p in match_any):
                return rule.get("score", 100), rule.get("field", "型号规则")
    return 0, ""


def search(keyword: str, max_results: int = 20) -> list:
    if not os.path.exists(DB_PATH):
        return []

    import openpyxl
    wb = openpyxl.load_workbook(DB_PATH, data_only=True, read_only=True)
    kw = keyword.lower().strip()
    is_type = _is_product_type(kw)
    results = []

    for sname in wb.sheetnames:
        if "恒流" in sname:
            continue
        ws = wb[sname]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if len(row) < 9:
                continue
            c3 = str(row[2] or "").strip()
            c4 = str(row[3] or "").strip()
            c5 = str(row[4] or "").strip() if len(row) > 4 else ""
            c8 = str(row[7] or "").strip() if len(row) > 7 else ""
            c9 = row[8] if len(row) > 8 else 0

            if not c3 and not c4:
                continue

            full = (c3 + " " + c4 + " " + c5).lower()

            if is_type and any(w in full for w in EXCLUDE_WORDS):
                continue

            score = 0
            field = ""

            score = 0
            field = ""

            if kw == c4.lower():
                score, field = 300, "型号精确"
            elif kw == c3.lower():
                score, field = 250, "货号精确"
            elif kw in c4.lower() and c4.lower().startswith(kw):
                score, field = 180, "名称包含"
            elif kw in sname.lower():
                if any(w in full for w in ["配件", "工具", "接头"]):
                    score, field = 30, "分类(附件)"
                else:
                    score, field = 150, "分类"
            elif kw in c4.lower():
                score, field = 120, "型号包含"
            elif kw in c3.lower():
                score, field = 90, "货号包含"
            elif kw in c5.lower():
                score, field = 30, "参数"

            # 关键词规则匹配（不干扰主链）
            if score == 0:
                r_score, r_field = _check_keyword_rules(c4, keyword, _keyword_rules)
                if r_score > 0:
                    score, field = r_score, r_field
            if score > 0:
                results.append({
                    "型号": c4 or c3, "货号": c3, "参数": c5,
                    "单位": c8, "单价": c9 or 0,
                    "sheet": sname, "score": score, "field": field,
                })

    wb.close()
    results.sort(key=lambda r: -r["score"])
    return results[:max_results]


def score_product(product: dict, keyword: str) -> dict:
    kw = keyword.lower().strip()
    model = (product.get("型号", "") or "").lower()
    parts_no = (product.get("货号", "") or "").lower()
    params = (product.get("参数", "") or "").lower()
    sheet = (product.get("sheet", "") or "").lower()
    is_type = _is_product_type(kw)

    full = model + " " + parts_no + " " + params
    if is_type and any(w in full for w in EXCLUDE_WORDS):
        product["score"] = -1
        product["field"] = "已排除"
        return product

    score = 0
    field = ""

    score = 0
    field = ""

    if kw == model:
        score, field = 300, "型号精确"
    elif kw == parts_no:
        score, field = 250, "货号精确"
    elif kw in model and model.startswith(kw):
        score, field = 180, "名称包含"
    elif kw in sheet:
        if any(w in full for w in ["配件", "工具", "接头"]):
            score, field = 30, "分类(附件)"
        else:
            score, field = 150, "分类"
    elif kw in model:
        score, field = 120, "型号包含"
    elif kw in parts_no:
        score, field = 90, "货号包含"
    elif kw in params:
        score, field = 30, "参数"

    # 关键词规则匹配（独立判断，不干扰主链）
    if score == 0:
        r_score, r_field = _check_keyword_rules(model, keyword, _keyword_rules)
        if r_score > 0:
            score, field = r_score, r_field

    product["score"] = score
    product["field"] = field
    return product


def score_with_params(product: dict, keyword: str, *,
                      color: str = None, cct: str = None,
                      beam: str = None, specs: str = None,
                      install: str = None) -> dict:
    """对产品评分，除基础 keyword 评分外，额外计算参数匹配分。

    评分规则（在基础分之上累加）：
      颜色匹配              +80
      色温匹配              +50
      光束角匹配            +50
      开孔/规格匹配         +30
      安装方式匹配          +30

    返回 product dict，score 为累加后总分。
    """
    import re
    if not product:
        return {}

    prod_copy = product.copy()
    result = score_product(prod_copy, keyword)
    base = result.get("score", 0)
    extra = 0
    reasons = []

    model = (product.get("型号", "") or "").lower()
    params_text = (product.get("参数", "") or "").lower()
    full_text = model + " " + params_text

    # 颜色匹配：检查产品型号/参数中是否包含颜色词
    if color:
        resolved = _resolve_color_keyword(color).lower()
        if resolved and resolved in full_text:
            extra += 80
            reasons.append("颜色+80")
        else:
            for alias_variant in _color_aliases.get(color, []):
                if alias_variant.lower() in full_text:
                    extra += 80
                    reasons.append("颜色+80")
                    break

    # 色温匹配
    if cct:
        cct_kw = cct.lower().replace("k", "").strip()
        if cct_kw in full_text:
            extra += 50
            reasons.append("色温+50")

    # 光束角匹配
    if beam:
        beam_kw = beam.lower().replace("°", "度").strip()
        if beam_kw in full_text:
            extra += 50
            reasons.append("光束角+50")

    # 开孔/规格匹配
    if specs:
        hole_num = re.search(r"(\d+)", specs)
        if hole_num:
            hole_str = hole_num.group(1)
            if hole_str in full_text:
                extra += 30
                reasons.append("规格+30")

    # 安装方式匹配
    if install:
        if install == "预埋":
            # 明确标注预埋 → +150
            if "安装方式：预埋" in params_text or "安装方式:预埋" in params_text:
                extra += 150
                reasons.append("安装+150(预埋式)")
            # 明确标注嵌入式（与用户需求冲突）→ -100
            elif "安装方式：嵌入式" in params_text or "安装方式:嵌入式" in params_text:
                extra -= 100
                reasons.append("安装-100(嵌入式)")
            # 兜底：字段中出现"预埋"
            else:
                if "预埋" in full_text:
                    extra += 30
                    reasons.append("安装+30(预埋)")
        else:
            # 其他安装方式的通用匹配
            install_kw = install.lower().strip()
            if install_kw in full_text:
                extra += 30
                reasons.append("安装+30")

    result["score"] = base + extra
    if reasons:
        existing = result.get("field") or ""
        result["field"] = (existing + " + " + " + ".join(reasons)) if existing else " + ".join(reasons)
    return result
