"""产品文本解析器 - 自然语言 -> 报价订单"""
import re, json, os
from services.quote_schema import empty_order

_PRODUCT_SERIES = {
    "柔光三代": {"series": "柔光III代", "model_hint": "TH09055"},
    "TH09055": {"series": "柔光III代", "model_hint": "TH09055"},
    "09055": {"series": "柔光III代", "model_hint": "TH09055"},
    "GS920": {"series": "GS920", "model_hint": "GS920"},
}

_COLORS = {"哑白": "白+白", "哑黑": "黑+黑", "白色": "白+白",
           "黑色": "黑+黑", "白+白": "白+白", "黑+黑": "黑+黑",
           "白+黑": "白+黑", "白+镍": "白+镍"}

_CCT = ["3000K", "3500K", "4000K", "双色温", "2700K", "6000K", "6500K"]
_BEAM = ["15度", "24度", "36度", "50度"]
_BEAM_SYMBOLS = ["15°", "24°", "36°", "50°"]

# 产品类别（不是具体型号），优先级低于具体系列
CATEGORY_SERIES = {
    "射灯", "筒灯", "磁吸", "灯带", "洗墙灯", "轨道灯",
    "面板灯", "壁灯", "吊灯", "吸顶灯", "格栅灯",
}

# 系列名称 → 可搜索型号关键词 反向映射
_SERIES_TO_HINT = {}
for _name, _info in _PRODUCT_SERIES.items():
    _s = _info["series"]
    _h = _info["model_hint"]
    if _s and _h:
        if _s not in _SERIES_TO_HINT or len(_name) > 3:
            _SERIES_TO_HINT[_s] = _h


_SERIES_CFG = None
def _get_series_config() -> dict:
    global _SERIES_CFG
    if _SERIES_CFG is not None:
        return _SERIES_CFG
    cfg_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "product_series.json")
    try:
        with open(cfg_path, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
        _SERIES_CFG = data
    except:
        _SERIES_CFG = {"series": []}
    # 启动诊断：打印所有已加载系列
    print("======== SERIES CONFIG ========")
    series_list = _SERIES_CFG.get("series", [])
    print(f"Loaded series count: {len(series_list)}")
    for _s in series_list:
        _name = _s.get("name", "")
        _patterns = _s.get("patterns", [])
        _models = _s.get("models", {})
        # 兼容旧结构
        if not _models and _s.get("standard"):
            _models = {
                "normal": _s.get("standard", ""),
                "normal_k2": (_s.get("variants", {}) or {}).get("K2", ""),
            }
            print(f"  [legacy] {_name}")
        print(f"  name: {_name}")
        print(f"  aliases: {_patterns}")
        # 兼容按开孔分组结构
        if isinstance(_models, dict) and any(isinstance(v, dict) for v in _models.values()):
            for _hk, _hm in _models.items():
                print(f"  {_hk}开孔: {_hm}")
        else:
            print(f"  normal: {_models.get('normal', '')}")
            print(f"  normal_k2: {_models.get('normal_k2', '')}")
            print(f"  pre_embedded: {_models.get('pre_embedded', '')}")
            print(f"  pre_embedded_k2: {_models.get('pre_embedded_k2', '')}")
        # 配置完整性检查
        _has_any = False
        if isinstance(_models, dict):
            for _v in _models.values():
                if isinstance(_v, dict) and any(_v.values()):
                    _has_any = True
                    break
                if isinstance(_v, str) and _v:
                    _has_any = True
                    break
        if not _has_any:
            print(f"  [CONFIG ERROR] {_name} 所有型号配置为空")
    print("===============================")
    return _SERIES_CFG


def _resolve_series_from_text(line: str) -> tuple:
    """从输入行解析系列。返回 (系列名, 占位型号, variant="")。
    
    实际型号在 parse_product_text 末尾根据安装方式（TH/WB）和色温（K2）决定。
    """
    cfg = _get_series_config()
    line_lower = line.lower()
    detected = []
    for s in cfg.get("series", []):
        _patterns = s.get("patterns", [])
        for pattern in _patterns:
            if pattern.lower() in line_lower:
                detected.append((s.get("name", ""), pattern, s))
                break
    if detected:
        # 优先级：具体系列 > 产品类别；同级别取最长别名匹配
        def _series_sort_key(item):
            s_name, pattern, s = item
            is_category = 1 if s_name in CATEGORY_SERIES else 0
            return (is_category, -len(pattern))
        detected.sort(key=_series_sort_key)
        s_name, matched_pattern, s = detected[0]
        models = s.get("models", {})
        if not models and s.get("standard"):
            models = {
                "normal": s.get("standard", ""),
                "normal_k2": (s.get("variants", {}) or {}).get("K2", ""),
            }
        base = models.get("normal", "") or s_name
        print(f"======== SERIES MATCH ========")
        print(f"Original text: {line[:50]}")
        print(f"Detected alias: {matched_pattern}")
        print(f"Candidate series: {[d[0] for d in detected]}")
        print(f"Selected series: {s_name}")
        print(f"Base model: {base}")
        print("==============================")
        return s_name, base, ""
    # 未匹配：打印可用别名
    print("======== SERIES MATCH ========")
    print(f"Original text: {line[:50]}")
    print("No series matched.")
    print("Available aliases:")
    _all = []
    for s in cfg.get("series", []):
        _all.extend(s.get("patterns", []))
    print("  " + " / ".join(_all))
    print("==============================")
    return "", "", ""


def _resolve_color(input_color):
    if not input_color:
        return ""
    rules_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "product_keyword_rules.json")
    try:
        with open(rules_path, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
            alias_map = data.get("color_alias", {})
            if input_color in alias_map:
                return alias_map[input_color][0]
    except:
        pass
    return _COLORS.get(input_color, input_color)


def get_series_model(series: str, install_type: str = "", is_k2: bool = False) -> str:
    """系列名 → 真实型号兜底。
    
    优先使用配置中的 QY- 型号；
    配置为系列名时，通过预埋盒配套关系查找对应灯具型号。
    """
    cfg = _get_series_config()
    for s in cfg.get("series", []):
        if s.get("name") == series:
            models = s.get("models", {})
            if install_type == "预埋":
                m = models.get("pre_embedded_k2" if is_k2 else "pre_embedded", "")
            else:
                m = models.get("normal_k2" if is_k2 else "normal", "")
            if m.startswith("QY-"):
                return m
            break
    # 通过预埋盒配套关系查找：参数含系列名的 ZH 盒 → 提取适配 WB 灯具
    import openpyxl
    db_path = r"D:\0000AAA报价助手\1.2026企一产品报价表(最新）.xlsx"
    try:
        wb = openpyxl.load_workbook(db_path, data_only=True)
    except Exception as e:
        print(f"[系列映射] 兜底查库失败: {e}")
        return ""
    try:
        for sname in wb.sheetnames:
            if "恒流" in sname:
                continue
            ws = wb[sname]
            for row_idx in range(3, (ws.max_row or 3) + 1):
                c4 = str(ws.cell(row=row_idx, column=4).value or "").strip()
                c5 = str(ws.cell(row=row_idx, column=5).value or "").strip()
                if c4.startswith("QY-ZH") and series in c5:
                    for m in re.findall(r"QY-WB[A-Za-z0-9\-]+", c5):
                        print(f"[系列映射] 兜底: {series} → {m}")
                        return m
                    for m in re.findall(r"WB[A-Za-z0-9\-]+", c5):
                        full = "QY-" + m
                        print(f"[系列映射] 兜底: {series} → {full}")
                        return full
    finally:
        wb.close()
    return ""


def _match_category_product(category: str, color: str = "", specs: str = "",
                            cct: str = "", beam: str = "", install_type: str = "") -> str:
    """产品类别（射灯/筒灯等）匹配真实型号。
    
    通过数据库评分搜索，结合颜色/色温/光束角/开孔/安装方式选出最佳型号。
    """
    from services.product_database_search_service import search as db_search, score_with_params
    candidates = db_search(category, max_results=20)
    if not candidates:
        print(f"[类别匹配] 无候选: {category}")
        return ""
    scored = []
    for c in candidates:
        s = score_with_params(
            c, category,
            color=color, cct=cct, beam=beam, specs=specs, install=install_type,
        )
        if s.get("score", 0) > 0:
            scored.append((s.get("score", 0), c))
    if not scored:
        print(f"[类别匹配] 无有效评分: {category}")
        return candidates[0].get("型号", "")
    scored.sort(key=lambda x: -x[0])
    best = scored[0][1]
    print(f"[类别匹配] {category} → {best.get('型号','')} (score={scored[0][0]})")
    return best.get("型号", "")


def _extract_qty(line: str) -> int:
    """从行中提取数量。支持：
      2套 / 2个 / 8米          → 数字+单位
      ×2 / x2 / X2             → 乘号+数字
      数量2 / 数量:2 / 数量：2  → 数量+数字
    """
    if not line:
        return 0
    m = re.search(r"(\d+)\s*[个套只盏米条根台]", line)
    if m:
        return int(m.group(1))
    m = re.search(r"[×xX]\s*(\d+)", line)
    if m:
        return int(m.group(1))
    m = re.search(r"数量\s*[:：]?\s*(\d+)", line)
    if m:
        return int(m.group(1))
    return 0


def _select_model_for_item(item: dict):
    """为单个订单项选择最终型号。
    
    匹配优先级：系列 → K2 → 安装方式(TH/WB) → 开孔尺寸 → 数据库型号。
    """
    cfg = _get_series_config()
    for s in cfg.get("series", []):
        if s.get("name") == item.get("_series_key", ""):
            models_all = s.get("models", {})
            # 提取开孔尺寸
            hole_size = ""
            hole_m = re.search(r"(\d+)", item.get("规格", ""))
            if hole_m:
                hole_size = hole_m.group(1)
            # 开孔优先：按开孔选择型号组
            models = models_all.get(hole_size, {}) if hole_size else {}
            if not models:
                # 兼容旧结构（无按开孔分组）
                models = models_all
            # 兼容旧结构 standard/variants
            if not models and s.get("standard"):
                models = {
                    "normal": s.get("standard", ""),
                    "normal_k2": (s.get("variants", {}) or {}).get("K2", ""),
                    "pre_embedded": s.get("standard", ""),
                    "pre_embedded_k2": (s.get("variants", {}) or {}).get("K2", ""),
                }
            pre = item.get("pre_embedded")
            dual = item.get("has_dual_cct")
            if pre:
                model = models.get("pre_embedded_k2") if dual else models.get("pre_embedded", "")
                label = "预埋(WB)"
            else:
                model = models.get("normal_k2") if dual else models.get("normal", "")
                label = "普通(TH)"
            if not model:
                model = item.get("_series_key", "")
            print(f"[型号选择] 系列: {item.get('_series_key','')}  开孔: {hole_size or '未指定'}  候选型号: {models}")
            print(f"[型号选择] 最终选择: {model}")
            item["_std_model"] = model
            item["_variant"] = "K2" if dual else ""
            print(f"[系列映射] {label} + {'双色温' if dual else '单色温'} → {model}")
            break


def parse_product_text(text):
    lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
    result = {"系列": "", "类型": "", "规格": "", "颜色": "", "安装": "", "参数": {}, "订单": []}
    current_cct = ""
    current_beam = ""
    current_product = None

    for line in lines:
        s_name, s_std, s_variant = _resolve_series_from_text(line)
        if s_std:
            # 新产品上下文：每行独立产品
            current_product = {
                "系列": s_name,
                "_series_key": s_name,
                "_std_model": s_std,
                "_variant": "",
                "安装": "",
                "颜色": "",
                "规格": "",
                "pre_embedded": False,
                "has_dual_cct": False,
            }
            result["系列"] = s_name
            result["_std_model"] = s_std
            result["_variant"] = s_variant
            result["_series_key"] = s_name
            # 新系列开始：重置上一产品的色温/光束角状态
            current_cct = ""
            current_beam = ""
            print(f"[系列映射] 输入行: {line[:30]}...  系列={s_name}  标准型号={s_std}  变体={s_variant}")
        if current_product is None:
            continue
        for name, info in _PRODUCT_SERIES.items():
            if name in line:
                result["系列"] = info["series"]
                break
        for cn, cv in _COLORS.items():
            raw = line.replace("杯", "").replace("色", "")
            if cn in raw:
                result["颜色"] = cv
                current_product["颜色"] = cv
                break
        m = re.search(r"(\d+)\s*开孔", line)
        if m:
            result["规格"] = m.group(1) + "开孔"
            result["参数"]["开孔"] = m.group(1)
            current_product["规格"] = m.group(1) + "开孔"
        if "预埋" in line or "嵌入" in line or "无边框" in line:
            result["安装"] = "预埋"
            current_product["安装"] = "预埋"
            current_product["pre_embedded"] = True
        elif "明装" in line:
            result["安装"] = "明装"
            current_product["安装"] = "明装"
        cct_found = ""
        for c in _CCT:
            if c in line:
                cct_found = c
                current_cct = c
                result["参数"]["色温"] = c
                if c == "双色温":
                    current_product["has_dual_cct"] = True
                break
        beam_found = ""
        for b in _BEAM:
            if b in line:
                beam_found = b
                break
        if not beam_found:
            for b in _BEAM_SYMBOLS:
                if b in line:
                    beam_found = b.replace("°", "度")
                    break
        if not beam_found:
            # 也支持不带单位的纯数字如 "24" 在 24° 上下文中
            m_beam = re.search(r"(\d+)\s*[°度]", line)
            if m_beam:
                deg = m_beam.group(1)
                beam_found = deg + "度"
        if beam_found:
            current_beam = beam_found
        # 数量可能在本行，也可能在下一行（如 "24度" 换行后 "2套"）
        qty = _extract_qty(line)
        if qty > 0:
            item = dict(current_product)
            item.update({"色温": current_cct, "光束角": current_beam, "数量": qty})
            _select_model_for_item(item)
            result["订单"].append(item)
    # 系列已识别但没有数量行时，生成默认订单项（数量=1）
    if not result["订单"] and current_product:
        item = dict(current_product)
        item.update({"色温": current_cct, "光束角": current_beam, "数量": 1})
        _select_model_for_item(item)
        result["订单"].append(item)
        print("[系列映射] 无数量行，默认生成 1 个订单项")
    return result


def build_order_from_parsed(parsed):
    from services.product_lookup_service import exact_db_lookup
    order = empty_order()
    for item in parsed.get("订单", []):
        series = item.get("系列", "")
        color = _resolve_color(item.get("颜色", ""))
        specs = item.get("规格", "")
        install_type = item.get("安装", "")
        is_k2 = item.get("_variant", "") == "K2"
        cct = item.get("色温", "")
        beam = item.get("光束角", "")
        # 每个订单项独立的最终型号（parse_product_text 已按项选择）
        std_model = item.get("_std_model", "")
        if not std_model:
            print(f"[LOOKUP ERROR] 未解析到最终型号，系列={series}")
            continue
        # 系列名兜底：非 QY- 型号时映射为真实型号
        if not std_model.startswith("QY-"):
            mapped = get_series_model(series, install_type, is_k2)
            if not mapped and series in CATEGORY_SERIES:
                mapped = _match_category_product(series, color, specs, cct, beam, install_type)
            if mapped:
                print(f"[系列映射] 系列名兜底: {std_model} -> {mapped}")
                std_model = mapped
        print(f"[系列映射] 最终型号: {std_model}")

        qty = item.get("数量", 0)
        # 构造完整型号（含颜色）
        search_model = std_model
        # 如果 std_model 已包含颜色后缀（有空格），不再拼接
        has_color_suffix = len(std_model.split()) > 1
        if color and not has_color_suffix:
            search_model = f"{std_model} {color}"
        # 精确查库：唯一数据库记录
        force_wb = (install_type == "预埋")
        record = exact_db_lookup(search_model, qty, force_wb=force_wb)
        # K2 变体不存在时，回退同系列基础型号（不跨系列）
        if not record and "-K2" in search_model:
            base_model = search_model.replace("-K2", "")
            print(f"[降级] K2 型号未找到，回退同系列基础型号: {base_model}")
            record = exact_db_lookup(base_model, qty, force_wb=force_wb)
        if not record:
            print(f"[LOOKUP ERROR] Model: {search_model}")
            continue
        # 开孔校验：预埋订单跳过；普通 TH 型号才校验
        record_model = record.get("型号", "")
        is_pre_embedded = (
            force_wb
            or install_type == "预埋"
            or search_model.startswith("QY-WB")
            or record_model.startswith("QY-WB")
        )
        if specs and not is_pre_embedded:
            hole_m = re.search(r"(\d+)", specs)
            record_hole = re.sub(r"[^0-9]", "", record.get("开孔", "") or "")
            if hole_m and record_hole and hole_m.group(1) not in record_hole:
                print(f"[开孔不匹配] 请求 {hole_m.group(1)}开孔，实际记录 {record.get('开孔','')}，跳过")
                continue
        print("======== EXACT LOOKUP ========")
        print(f"Model: {search_model}")
        print(f"Found: {record.get('型号','')}")
        print(f"Price: {record.get('单价',0)}")
        print(f"Sheet: {record.get('db_sheet','')}")
        print(f"Row: {record.get('db_row','')}")
        print("==============================")
        price = record.get("单价", 0) or 0
        order["products"].append({
            "model": record.get("型号", search_model),
            "name": record.get("型号", ""),
            "color": color,
            "cct": cct,
            "beam": beam,
            "_series_key": series,
            "pre_embedded": force_wb,
            "quantity": qty,
            "unit_price": price,
            "total_price": price * qty,
            "_db_record": record,
        })
    return order
